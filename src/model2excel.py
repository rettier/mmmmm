import numpy as np
from keras.engine.saving import load_model
from openpyxl import Workbook, styles
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

input_keys_template = ['max', 'min', 'sum', 'sum_without_max', 'avg', 'avg_without_max', 'rows']
input_keys = None
features = 7
hidden = 5
outputs = 11


def to_igschel_hiso(model, name):
    input_keys = input_keys_template
    if features == 6:
        del input_keys[1]

    wb = Workbook()
    ws = wb.active

    for i, x in enumerate(range(features)):
        ws.cell(row=i + 1, column=1).value = input_keys[i]
        ws.cell(row=i + 1, column=2).value = 1

    for x in range(hidden):
        cell = chr(ord("A") + x)
        ws.cell(row=x + 1, column=4).value = \
            "=SUMPRODUCT(B1:B{features}, weights0!{cell}1:{cell}{features}) + weights1!A{idx}" \
                .format(cell=cell, features=features, idx=x + 1)

    for x in range(outputs):
        cell = chr(ord("A") + x)
        ws.cell(row=x + 1, column=6).value = 2 ** x
        ws.cell(row=x + 1, column=7).value = \
            "=SUMPRODUCT(D1:D{hidden}, weights2!{cell}1:{cell}{hidden}) + weights3!A{idx}" \
                .format(cell=cell, hidden=hidden, idx=x + 1)

    rule = ColorScaleRule(start_type='percentile', start_value=40, start_color='ffb6d7a8',
                          mid_type='percentile', mid_value=70, mid_color='ff9fc5e8',
                          end_type='percentile', end_value=95, end_color='ffea9999')

    bold = styles.Font(bold=True)

    ws.conditional_formatting.add('D1:D{}'.format(hidden), rule)
    ws.conditional_formatting.add('G1:G{}'.format(outputs), rule)
    ws.conditional_formatting.add('G1:G{}'.format(outputs),
                                  CellIsRule(operator='equal', formula=['MAX(G$1:G${})'.format(outputs)], font=bold))

    w = model.get_weights()
    for y in range(4):
        weights = wb.create_sheet("weights{}".format(y))
        wy = w[y]
        shape = wy.shape
        if len(shape) == 2:
            for ix, iy in np.ndindex(shape):
                weights.cell(row=ix + 1, column=iy + 1).value = float(wy[ix, iy])
        else:
            for ix in range(shape[0]):
                weights.cell(row=ix + 1, column=1).value = float(wy[ix])

    wb.save(name + ".xlsx")


to_igschel_hiso(load_model("models/7x5x11-full.h5"), "7")

features = 6
to_igschel_hiso(load_model("models/6x5x11-full.h5"), "6")
